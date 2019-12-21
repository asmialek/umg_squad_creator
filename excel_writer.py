import openpyxl


def write_to_excel(mech):
    wb = openpyxl.load_workbook('template.xlsx')
    ws = wb['Sheet1']
    # ws['A1'] = 'A1'
    # wb.save('template_used.xlsx')

    cells = {
            'F2': 'name',
            'F4': 'frame_name',
            # 'M5': 'symbol',
            'M7': 'pts',
            'M9': 'MV',
            'M11': 'EG',
            'O6': 'AM',
            'K14': 'special'
            }

    for cell in cells:
        print(cell, cells[cell], getattr(mech, cells[cell]))
        ws[cell] = getattr(mech, cells[cell])

    letters_odd = ['B', 'B', 'G', 'H', 'I', 'J', 'B']
    letters_even = ['K', 'K', 'P', 'Q', 'R', 'S', 'K']
    numbers = [16, 17, 17, 17, 17, 17, 18]
    params = ['name', 'module_type', 'EC', 'DM', 'RG', 'pts', 'special']
    for k in range(2):
        if k == 0:
            letters = letters_odd
        else:
            letters = letters_even
        for j in range(3):
            for i in range(7):
                cell = letters[i] + str(numbers[i] + j*5)
                print(list(mech.slots.keys()), len(list(mech.slots.keys())))
                if len(list(mech.slots.keys())) > j*2+k:
                    slot = mech.slots[list(mech.slots.keys())[j*2+k]]
                    if slot:
                        print(cell, getattr(slot, params[i]))
                        ws[cell] = getattr(slot, params[i])
    img = openpyxl.drawing.image.Image('justin-spice-beta-mech-thumbs.jpg')
    img.anchor = 'B5'
    img.width = 286
    img.height = 319
    ws.add_image(img)

    wb.save('template_mechs_new.xlsx')


#
# 'B16 = name
# 'G17 = EC
# 'H17 = DM
# 'I17 = RG
# 'J17 = pts
# 'B18 = special
#
# K,P,Q,R
# K
#
# 21
# 22
# 23
#
# 26
# 27
# 28
